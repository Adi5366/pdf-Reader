import pandas as pd
from openpyxl import load_workbook

# Your dictionary
my_dict = {'Jan-24': [1, 2, 3, 4], 'Feb-24': [3, 4, 5, 6], 'Mar-24': [4, 5, 6, 7]}

# Create a DataFrame
df = pd.DataFrame(my_dict)

# Specify the filename
filename = 'output.xlsx'

# Save the DataFrame to an Excel file
df.to_excel(filename, index=False, sheet_name='Data')

# Load the workbook and select the active worksheet
wb = load_workbook(filename)
ws = wb['Data']

# Define the starting cell
start_row = 19
start_col = 13  # Column M is the 13th column

# Write the DataFrame to the specified cell
for r_idx, row in enumerate(df.values, start=start_row):
    for c_idx, value in enumerate(row, start=start_col):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Save the workbook
wb.save(filename)
